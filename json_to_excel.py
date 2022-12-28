from importlib.resources import path
from openpyxl import Workbook, load_workbook
import os
from helpful_scripts import *
from file_handler import *
import constants
from datetime import date


def json_to_excel():
    today = str(date.today())
    excel_path = f"./data/{today}/{today}.xlsx"
    try:

        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        for brand in constants.BRANDS:

            data_file = get_today_file_path(f"{brand}.json")
            with open(data_file) as f:
                items = json.load(f)

            ws = wb.create_sheet(brand)

            for index, i in enumerate(items[0].keys()):
                header_cell = ws.cell(row=1, column=index + 1)
                header_cell.value = i

            for i in items:
                item = [
                    i["Name"],
                    i["Description"],
                    i["Price"],
                    i["Location"],
                    i["Views"],
                    i["Post Date"],
                    i["User Name"],
                    i["Telephone"],
                    i["Post Link"],
                    i["Image"],
                ]
                ws.append(item)

        wb.save(excel_path)
    except Exception:
        print(f"Failed to crea excel file: {path}")


if __name__ == "__main__":
    json_to_excel()
